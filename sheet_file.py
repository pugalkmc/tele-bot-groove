import os

import openpyxl

from functions import *


async def spreadsheet(chat_id, task_id=None, date=None):
    if date is None:
        collection_name = datetime.datetime.now().strftime("%d-%m-%Y")
    else:
        collection_name = date
    wb = openpyxl.Workbook()
    ws = wb.active
    # Write the headers
    task_details = tasks_col.find_one({"task_id": int(task_id)})
    if not task_details:
        await bot.send_message(chat_id=chat_id, text="Task id not valid!")
        return 0
    task_type = task_details['task_type']
    ws.column_dimensions['A'].width = 14
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 18

    ws['F1'] = 'Username'
    ws['G1'] = 'Count'
    flex_title = "Message text" if task_type == "normal" else "filter"

    ws['A1'] = 'Username'
    ws['B1'] = flex_title
    ws['C1'] = 'IST Time'

    username_counts = {}
    if 'collection' in task_details and collection_name in task_details['collection']:
        row = 2
        for task_info in task_details['collection'][collection_name]:
            if task_info is not None and isinstance(task_info, dict):
                username = task_info.get('username')
                text = task_info.get('text')
                time = task_info.get('time')
                if username is not None:
                    ws.cell(row=row, column=1).value = username
                if text is not None:
                    ws.cell(row=row, column=2).value = text
                if time is not None:
                    ws.cell(row=row, column=3).value = time

                if username in username_counts:
                    username_counts[username]['count'] += 1
                else:
                    username_counts[username] = {'count': 1}
                row += 1
    row = 1  # Assuming the initial row is 1
    for username, data in username_counts.items():
        count = data['count']
        ws.cell(row=row, column=6).value = username
        ws.cell(row=row, column=7).value = count
        row += 1

    wb.save(f"{collection_name}_{task_id}.xlsx")
    await bot.send_document(chat_id=chat_id, document=open(f"{collection_name}_{task_id}.xlsx", "rb"))
    wb.close()
    os.remove(f"{collection_name}_{task_id}.xlsx")
