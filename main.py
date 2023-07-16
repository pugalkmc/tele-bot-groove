import os
import re
import openpyxl
from telegram.ext import CommandHandler, MessageHandler, filters, Application
from edit_task import *
from settings_con import *
from task_create import *


def get_ip_address():
    url = "https://checkip.amazonaws.com/"
    payload = {}
    headers = {}
    response = requests.request("GET", url, headers=headers, data=payload)
    return response.text


client = Client(api_key, api_secret)


async def start(update, context):
    message = update.message
    chat_id = message.chat_id
    username = message.from_user.username
    if message.chat.type in ['group', 'supergroup']:
        return
    current_time = time_fun.now().strftime("%d-%m-%Y")

    find_people = peoples_col.find_one({"chat_id": chat_id})
    if username is None:
        await bot.send_message(chat_id=chat_id, text=f"Hello @{username}\n"
                                                     f"Please set your telegram username in settings!\n"
                                                     f"Check out this document for guidance: Not set")
    elif find_people is None:
        peoples_col.insert_one({"chat_id": chat_id, "username": username, "first_started": current_time})
        await bot.send_message(chat_id=chat_id, text=f"Hello! welcome @{username}")
    else:
        await bot.send_message(chat_id=chat_id, text="Hi! welcome back")
    await menu_button(update, context)


async def cancel(update, context):
    message = update.message
    chat_id = message.chat_id
    if message.chat.type == 'group' or message.chat.type == 'supergroup':
        return
    reply_keyboard = [["settings", "My task"]]
    await bot.send_message(chat_id=chat_id, text="Use menu buttons for quick access",
                           reply_markup=ReplyKeyboardMarkup(reply_keyboard, resize_keyboard=True,
                                                            one_time_keyboard=True),
                           reply_to_message_id=update.message.message_id)
    return ConversationHandler.END


async def private_message_handler(update, context):
    message = update.message
    chat_id = message.chat_id
    text = message.text.lower()
    username = update.message.chat.username.lower()
    if username is None:
        await bot.send_message(chat_id=chat_id, text=f"Hello @{username}\n"
                                                     f"Please set your telegram username in settings!\n"
                                                     f"Check out this document for guidance: Not set")
        return
    if text == "cancel":
        await cancel(update, context)
    if "ip address" == text:
        ip_address = get_ip_address()
        await bot.send_message(chat_id=chat_id, text=ip_address)
    elif "pay " in text:
        details_pay = text.split(" ")
        await functions.binance_pay(details_pay[1], details_pay[2], details_pay[3], chat_id)

    elif "my task" == text:
        count = 0
        text = ""
        each_task = tasks_col.find({})
        for each in each_task:
            if await verify_membership(update, each):
                count += 1
                text += f"Title : {each['title']}\nCommand: <code>sheet {each['task_id']}</code>\n\n"
        if count == 0:
            await bot.send_message(chat_id=chat_id, text="You are not assigned for any task right now!")
        else:
            await bot.send_message(chat_id=chat_id, text=f"Total found: {count}\n\n"
                                                         f"{text}"
                                                         f"Just click to copy the command and send it here",
                                   parse_mode="html")
    elif "sheet " in text:
        task_id = text.split(" ")
        if not task_id[1].isnumeric():
            await bot.send_message(chat_id=chat_id, text=f"task id must be number\n")
        elif len(task_id) == 2:
            await sheet_file.spreadsheet(chat_id, task_id[1])
        else:
            await sheet_file.spreadsheet(chat_id, task_id[1], task_id[2])
    elif "payment data" == text:
        query_data = {
            '$or': [
                {'$and': [{'binance': {'$exists': True}}, {'upi': {'$exists': True}}]},
                {'binance': {'$exists': True}, 'upi': {'$exists': False}},
                {'binance': {'$exists': False}, 'upi': {'$exists': True}},
                {'binance': {'$exists': False}, 'upi': {'$exists': False}}
            ]
        }
        get = peoples_col.find(query_data)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 30
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 20
        ws['A1'] = 'Username'
        ws['B1'] = 'TUSD Address'
        ws['C1'] = 'Binance'
        ws['D1'] = 'UPI'
        row = 2
        for data in get:
            if 'username' not in data:
                continue
            username = data['username']
            ws.cell(row=row, column=1).value = username
            ws.cell(row=row, column=2).value = data['address'] if 'address' in data else "Not set"
            ws.cell(row=row, column=3).value = data['binance'] if 'binance' in data else "Not set"
            ws.cell(row=row, column=4).value = data['upi'] if 'upi' in data else "Not set"
            row += 1
        wb.save(f"peoples_data.xlsx")
        await bot.send_document(chat_id=chat_id, document=open(f"peoples_data.xlsx", "rb"))
        wb.close()
        os.remove("peoples_data.xlsx")

    elif "pause " in text:
        task_id = text.replace("pause ", "")
        await task_status_switch(chat_id, task_id, option="pause")

    elif "active " in text:
        task_id = text.replace("active ", "")
        await task_status_switch(chat_id, task_id, option="active")


async def group_message_handler(update, context):
    message = update.message
    text = message.text
    username = message.from_user.username
    user_id = message.from_user.id
    message_id = message.message_id
    group_id = message.chat.id

    collection_name = time_fun.now().strftime("%d-%m-%Y")
    message_date_ist = time_fun.now().strftime("%H:%M:%S")
    task = tasks_col.find_one({"work_group": group_id})
    if not task or task['status'] == 'paused':
        return
    if task['task_type'] == 'filter':
        if 'collection' in task and int(task['daily_target']) <= len(task['collection'][collection_name]):
            return
        elif not text or task['filter_text'] not in text:
            return
    else:
        if 'collection' in task and int(task['daily_target']) <= len(task['collection'][collection_name]):
            return
    if not await verify_membership(update, task):
        return
    new_message = {
        'username': username,
        'text': text,
        'message_id': message_id,
        'chat_id': message.from_user.id,
        'time': message_date_ist
    }

    tasks_col.update_one(
        {'task_id': task['task_id']},
        {"$push": {f'collection.{collection_name}': new_message}}
    )

    get = tasks_col.find_one({'task_id': task['task_id']})['collection'][collection_name]
    count = 0
    for user in get:
        count += 1 if user['username'] == username else 0
    if count == task['user_target']:
        flex_text = task['group_title'] if task['task_type'] == 'filter' else task['work_group_title']
        flex_text2 = "Stop your work for today" if task[
                                                       'task_type'] == 'filter' else "Still you can do more, if you like"
        await bot.send_message(chat_id=message.from_user.id,
                               text=f"Today target reached : {flex_text}\n"
                                    f"{flex_text2}")
        if task['task_type'] == 'filter':
            await bot.send_message(chat_id=task['group_id'],
                                   text=f"Today target reached : {flex_text}\n"
                                        f"{flex_text2}")
    if len(get) == task['daily_target']:
        if task['task_type'] == 'filter':
            await bot.send_message(chat_id=task['group_id'],
                                   text=f"Today target reached!\n\n"
                                        f"Stop your work , No more new records will be accepted")


async def task_status_switch(chat_id, task_id, option):
    if task_id.isnumeric():
        check = tasks_col.find_one({"task_id": task_id})
        if check is None:
            await bot.send_message(chat_id=chat_id, text="Invalid task ID")
            return
        get = tasks_col.update_one({"task_id": task_id}, {"$set": {"status": option}})
        await bot.send_message(chat_id=chat_id, text=f"{get['group_id']} : Task {option}!")
        await bot.send_message(chat_id=get['group_id'], text=f"Task {option}")
    else:
        await bot.send_message(chat_id=chat_id, text="Task id must be a number")


async def verify_membership(update, task):
    user_id = update.message.from_user.id
    try:
        member = await bot.get_chat_member(task['group_id'], user_id)
        if member.status in ['member', 'creator', 'administrator']:
            return True
        else:
            return False
    except Exception as e:
        return False


def main():
    dp = Application.builder().token(BOT_TOKEN).build()
    create = ConversationHandler(
        entry_points=[CommandHandler('create_task', create_task)],
        states={
            TITLE: [MessageHandler(filters.TEXT, title)],
            TASK_TYPE: [MessageHandler(filters.TEXT, task_type)],
            GROUP_ID: [MessageHandler(filters.TEXT, group_id)],
            FLEXIBLE_INPUT: [MessageHandler(filters.TEXT, flex_input)],
            USER_TARGET: [MessageHandler(filters.TEXT, user_target)],
            DAILY_TARGET: [MessageHandler(filters.TEXT, daily_target)],
            ADMINS_LIST: [MessageHandler(filters.TEXT, admins_list)],
            CONFIRM: [MessageHandler(filters.TEXT, confirm)]
        }, fallbacks=[]
    )

    twitter_settings = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(re.compile(r'^twitter$', re.IGNORECASE)), twitter_ids)],
        states={
            TWITTER_UPDATE: [MessageHandler(filters.TEXT, twitter_update)],
            TWITTER_UPDATE_LIST: [MessageHandler(filters.TEXT, twitter_update_list)],
            TWITTER_UPDATE_CONFIRM: [MessageHandler(filters.TEXT, twitter_update_confirm)],
        }, fallbacks=[]
    )

    binance = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(re.compile(r'^binance$', re.IGNORECASE)), binance_start)],
        states={
            BINANCE_OPTIONS: [MessageHandler(filters.TEXT, binance_option)],
            SET_BINANCE: [MessageHandler(filters.TEXT, set_binance)],
        }, fallbacks=[]
    )

    address = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(re.compile(r'^tusd address$', re.IGNORECASE)), address_start)],
        states={
            ADDRESS_OPTIONS: [MessageHandler(filters.TEXT, address_option)],
            SET_ADDRESS: [MessageHandler(filters.TEXT, set_address)],
        }, fallbacks=[]
    )

    discord = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(re.compile(r'^discord$', re.IGNORECASE)), discord_start)],
        states={
            DISCORD_OPTIONS: [MessageHandler(filters.TEXT, discord_option)],
            SET_DISCORD: [MessageHandler(filters.TEXT, set_discord)],
        }, fallbacks=[]
    )

    upi = ConversationHandler(
        entry_points=[MessageHandler(filters.TEXT & filters.Regex(re.compile(r'^upi id$', re.IGNORECASE)), upi_start)],
        states={
            UPI_OPTIONS: [MessageHandler(filters.TEXT, upi_option)],
            SET_UPI: [MessageHandler(filters.TEXT, set_upi)],
        }, fallbacks=[]
    )

    edit_task_con = ConversationHandler(
        entry_points=[
            MessageHandler(filters.TEXT & filters.Regex(re.compile(r'^admin mode$', re.IGNORECASE)), task_id)],
        states={
            ADMIN_MODE: [MessageHandler(filters.TEXT, admin_mode)],
            MEMBER_START: [MessageHandler(filters.TEXT, member_start)],
            SHEET_OPTION: [MessageHandler(filters.TEXT, sheet_option)],
            DATE_RANGE: [MessageHandler(filters.TEXT, date_range)],
            PROCEED_PAYMENT: [MessageHandler(filters.TEXT, proceed_payment)],
            PAYMENT_FOR_TASK: [MessageHandler(filters.TEXT, payment_for_task)],
            CHECK_PASSWORD: [MessageHandler(filters.TEXT, check_password)],
        }, fallbacks=[]
    )
    dp.add_handler(edit_task_con)
    dp.add_handler(upi)
    dp.add_handler(create)
    dp.add_handler(binance)
    dp.add_handler(address)
    dp.add_handler(discord)
    dp.add_handler(twitter_settings)
    dp.add_handler(CommandHandler("cancel", cancel))
    dp.add_handler(MessageHandler(filters.Regex(re.compile(r'^settings$', re.IGNORECASE)), settings))
    dp.add_handler(MessageHandler(filters.Regex(re.compile(r'^main menu$', re.IGNORECASE)), menu_button))
    dp.add_handler(CommandHandler("start", start))
    dp.add_handler(MessageHandler(filters.ChatType.PRIVATE, private_message_handler))
    dp.add_handler(MessageHandler(filters.ChatType.GROUPS, group_message_handler))
    dp.run_polling()


main()
